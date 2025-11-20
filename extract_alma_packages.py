#!/usr/bin/env python3
"""
extract_alma_packages.py

Read Alma E-Resource Activation Form (xlsm) and export all packages
from LICENSED_AGG, SELECTIVE_PKG, and DATABASE sheets into a CSV.

Usage:
    python extract_alma_packages.py "E-Resource Activation Form.xlsm" outputs/alma_packages.csv
"""

import sys
import csv
from openpyxl import load_workbook

SHEETS = ["LICENSED_AGG", "SELECTIVE_PKG", "DATABASE"]

def normalize(s: str) -> str:
    if s is None:
        return ""
    s = s.strip().lower()
    s = s.replace("&", "and")
    while "  " in s:
        s = s.replace("  ", " ")
    return s

def extract_sheet(ws, sheet_name, writer):
    # Expect header in row 1
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header = [cell.value for cell in header_row]

    try:
        pkg_id_idx = header.index("PACKAGE_ID")
        pkg_name_idx = header.index("PACKAGE_NAME")
        service_type_idx = header.index("SERVICE_TYPE")
    except ValueError as e:
        raise SystemExit(f"ERROR: Expected columns not found on sheet {sheet_name}: {e}")

    empty_streak = 0
    rows_written = 0

    # Use enumerate so we always know the Excel row number,
    # without relying on cell.row (which can be missing on EmptyCell).
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Check if the whole row is empty
        values = [cell.value for cell in row]
        if all(v is None for v in values):
            empty_streak += 1
            if empty_streak >= 20:
                # assume we're past the real data
                break
            continue
        empty_streak = 0

        pkg_id_cell = row[pkg_id_idx]
        pkg_name_cell = row[pkg_name_idx]
        service_type_cell = row[service_type_idx]

        pkg_id = pkg_id_cell.value
        pkg_name = pkg_name_cell.value

        # If both ID and name are empty, treat as effectively empty
        if pkg_id is None and pkg_name is None:
            continue

        service_type = service_type_cell.value
        norm_name = normalize(str(pkg_name) if pkg_name is not None else "")

        writer.writerow({
            "SheetName": sheet_name,
            "ExcelRow": row_idx,  # Excel's 1-based row number
            "Alma_PACKAGE_ID": str(pkg_id) if pkg_id is not None else "",
            "Alma_PACKAGE_NAME": str(pkg_name) if pkg_name is not None else "",
            "Alma_PACKAGE_NAME_norm": norm_name,
            "SERVICE_TYPE": str(service_type) if service_type is not None else "",
        })
        rows_written += 1

    print(f"Sheet {sheet_name}: wrote {rows_written} package rows")

def main(alma_path: str, out_path: str) -> None:
    # read_only=True is fine; we’re just scanning values
    wb = load_workbook(alma_path, data_only=False, read_only=True)

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "SheetName",
            "ExcelRow",
            "Alma_PACKAGE_ID",
            "Alma_PACKAGE_NAME",
            "Alma_PACKAGE_NAME_norm",
            "SERVICE_TYPE",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for sheet_name in SHEETS:
            if sheet_name not in wb.sheetnames:
                print(f"WARNING: Sheet {sheet_name} not found in workbook; skipping")
                continue
            ws = wb[sheet_name]
            extract_sheet(ws, sheet_name, writer)

    print(f"Wrote Alma packages to {out_path}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print('Usage: python extract_alma_packages.py "E-Resource Activation Form.xlsm" outputs/alma_packages.csv')
        raise SystemExit(1)
    main(sys.argv[1], sys.argv[2])
